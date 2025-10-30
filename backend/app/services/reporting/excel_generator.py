# Excel Report Generator - Phase B Sprint 2 Day 1
# Professional Excel report generation with charts and multi-sheet support

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

from typing import Dict, List, Any, Optional
from datetime import datetime
import json
import pandas as pd
import io

class ExcelReportGenerator:
    """
    Professional Excel report generation service with charts and formatting
    """
    
    def __init__(self):
        self.setup_styles()
    
    def setup_styles(self):
        """Setup custom cell styles for professional reports"""
        # Header style
        self.header_style = NamedStyle(name="header_style")
        self.header_style.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        self.header_style.fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
        self.header_style.alignment = Alignment(horizontal='center', vertical='center')
        self.header_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Title style
        self.title_style = NamedStyle(name="title_style")
        self.title_style.font = Font(name='Calibri', size=18, bold=True, color='1976D2')
        self.title_style.alignment = Alignment(horizontal='center', vertical='center')
        
        # KPI style
        self.kpi_style = NamedStyle(name="kpi_style")
        self.kpi_style.font = Font(name='Calibri', size=16, bold=True, color='4CAF50')
        self.kpi_style.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data style
        self.data_style = NamedStyle(name="data_style")
        self.data_style.font = Font(name='Calibri', size=11)
        self.data_style.alignment = Alignment(horizontal='left', vertical='center')
        self.data_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    def generate_report(self, template_data: Dict[str, Any], 
                       report_data: Dict[str, Any], 
                       output_path: str) -> Dict[str, Any]:
        """
        Generate an Excel report based on template and data
        
        Args:
            template_data: Report template configuration
            report_data: Data to populate the report
            output_path: Output file path
            
        Returns:
            Dict with generation results and metadata
        """
        start_time = datetime.now()
        
        try:
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Add custom styles to workbook
            try:
                wb.add_named_style(self.header_style)
                wb.add_named_style(self.title_style)
                wb.add_named_style(self.kpi_style)
                wb.add_named_style(self.data_style)
            except ValueError:
                # Styles already exist
                pass
            
            # Build sheets based on template configuration
            sheets_config = template_data.get('sheets', [])
            if not sheets_config:
                # Default single sheet configuration
                sheets_config = [{'name': 'Report', 'type': 'summary'}]
            
            for sheet_config in sheets_config:
                self._build_sheet(wb, sheet_config, report_data)
            
            # Save workbook
            wb.save(output_path)
            
            # Calculate generation time
            generation_time = (datetime.now() - start_time).total_seconds() * 1000
            
            # Get file size
            import os
            file_size = os.path.getsize(output_path)
            
            return {
                'success': True,
                'file_path': output_path,
                'file_size': file_size,
                'generation_time_ms': int(generation_time),
                'sheet_count': len(sheets_config),
                'format': 'xlsx'
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'generation_time_ms': int((datetime.now() - start_time).total_seconds() * 1000)
            }
    
    def _build_sheet(self, workbook: Workbook, sheet_config: Dict, report_data: Dict):
        """Build a worksheet based on configuration"""
        sheet_name = sheet_config.get('name', 'Sheet1')
        sheet_type = sheet_config.get('type', 'summary')
        
        # Create worksheet
        ws = workbook.create_sheet(title=sheet_name)
        
        # Add report title
        if report_data.get('title'):
            ws['A1'] = report_data['title']
            ws['A1'].style = self.title_style
            ws.merge_cells('A1:F1')
            current_row = 3
        else:
            current_row = 1
        
        # Add metadata section
        if sheet_config.get('show_metadata', True):
            current_row = self._add_metadata_section(ws, report_data.get('metadata', {}), current_row)
        
        # Build sheet content based on type
        if sheet_type == 'summary':
            self._build_summary_sheet(ws, report_data, current_row)
        elif sheet_type == 'department_breakdown':
            self._build_department_breakdown_sheet(ws, report_data, current_row)
        elif sheet_type == 'user_details':
            self._build_user_details_sheet(ws, report_data, current_row)
        elif sheet_type == 'overdue_analysis':
            self._build_overdue_analysis_sheet(ws, report_data, current_row)
        elif sheet_type == 'data_table':
            self._build_data_table_sheet(ws, sheet_config, report_data, current_row)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _add_metadata_section(self, worksheet, metadata: Dict, start_row: int) -> int:
        """Add metadata information section"""
        if not metadata:
            return start_row
        
        # Section title
        worksheet[f'A{start_row}'] = 'Report Information'
        worksheet[f'A{start_row}'].font = Font(size=12, bold=True)
        start_row += 1
        
        # Metadata items
        if metadata.get('generated_at'):
            worksheet[f'A{start_row}'] = 'Generated:'
            worksheet[f'B{start_row}'] = metadata['generated_at']
            start_row += 1
        
        if metadata.get('generated_by'):
            worksheet[f'A{start_row}'] = 'Generated by:'
            worksheet[f'B{start_row}'] = metadata['generated_by']
            start_row += 1
        
        if metadata.get('department'):
            worksheet[f'A{start_row}'] = 'Department:'
            worksheet[f'B{start_row}'] = metadata['department']
            start_row += 1
        
        if metadata.get('period'):
            worksheet[f'A{start_row}'] = 'Period:'
            worksheet[f'B{start_row}'] = metadata['period']
            start_row += 1
        
        return start_row + 2
    
    def _build_summary_sheet(self, worksheet, report_data: Dict, start_row: int):
        """Build summary sheet with KPIs and overview data"""
        current_row = start_row
        
        # KPIs section
        kpis = report_data.get('kpis', [])
        if kpis:
            worksheet[f'A{current_row}'] = 'Key Performance Indicators'
            worksheet[f'A{current_row}'].font = Font(size=14, bold=True)
            current_row += 2
            
            # KPI headers
            headers = ['Metric', 'Current Value', 'Previous Value', 'Target', 'Status', 'Trend']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=current_row, column=col, value=header)
                cell.style = self.header_style
            
            current_row += 1
            
            # KPI data
            for kpi in kpis:
                row_data = [
                    kpi.get('name', ''),
                    f"{kpi.get('value', 0)}{kpi.get('unit', '')}",
                    f"{kpi.get('previous_value', 0)}{kpi.get('unit', '')}",
                    f"{kpi.get('target', 'N/A')}{kpi.get('unit', '') if kpi.get('target') else ''}",
                    kpi.get('status', ''),
                    kpi.get('trend', '')
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=current_row, column=col, value=value)
                    cell.style = self.data_style
                    
                    # Color code status
                    if col == 5:  # Status column
                        if value.lower() == 'excellent':
                            cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
                        elif value.lower() == 'good':
                            cell.fill = PatternFill(start_color='DCEDC8', end_color='DCEDC8', fill_type='solid')
                        elif value.lower() == 'warning':
                            cell.fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
                        elif value.lower() == 'critical':
                            cell.fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
                
                current_row += 1
            
            current_row += 2
        
        # Summary statistics
        summary = report_data.get('summary', {})
        if summary:
            worksheet[f'A{current_row}'] = 'Summary Statistics'
            worksheet[f'A{current_row}'].font = Font(size=14, bold=True)
            current_row += 2
            
            for key, value in summary.items():
                worksheet[f'A{current_row}'] = key.replace('_', ' ').title()
                worksheet[f'B{current_row}'] = value
                current_row += 1
    
    def _build_department_breakdown_sheet(self, worksheet, report_data: Dict, start_row: int):
        """Build department breakdown analysis sheet"""
        current_row = start_row
        
        department_data = report_data.get('departments', [])
        if not department_data:
            worksheet[f'A{current_row}'] = 'No department data available'
            return
        
        # Department breakdown title
        worksheet[f'A{current_row}'] = 'Department Performance Breakdown'
        worksheet[f'A{current_row}'].font = Font(size=14, bold=True)
        current_row += 2
        
        # Headers
        headers = ['Department', 'Quality Score', 'Training Completion', 'Document Efficiency', 'Overall Rating']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=current_row, column=col, value=header)
            cell.style = self.header_style
        
        current_row += 1
        
        # Department data
        for dept in department_data:
            row_data = [
                dept.get('name', ''),
                f"{dept.get('quality_score', 0):.1f}%",
                f"{dept.get('training_completion', 0):.1f}%",
                f"{dept.get('document_efficiency', 0):.1f} days",
                dept.get('overall_rating', 'N/A')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=current_row, column=col, value=value)
                cell.style = self.data_style
            
            current_row += 1
        
        # Add chart if we have data
        if len(department_data) > 1:
            self._add_department_chart(worksheet, current_row + 2, len(department_data))
    
    def _build_user_details_sheet(self, worksheet, report_data: Dict, start_row: int):
        """Build individual user details sheet"""
        current_row = start_row
        
        user_data = report_data.get('users', [])
        if not user_data:
            worksheet[f'A{current_row}'] = 'No user data available'
            return
        
        # User details title
        worksheet[f'A{current_row}'] = 'Individual User Status'
        worksheet[f'A{current_row}'].font = Font(size=14, bold=True)
        current_row += 2
        
        # Headers
        headers = ['Name', 'Department', 'Training Status', 'Completion %', 'Overdue Items', 'Last Activity']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=current_row, column=col, value=header)
            cell.style = self.header_style
        
        current_row += 1
        
        # User data
        for user in user_data:
            row_data = [
                user.get('name', ''),
                user.get('department', ''),
                user.get('training_status', ''),
                f"{user.get('completion_percentage', 0):.1f}%",
                user.get('overdue_items', 0),
                user.get('last_activity', 'N/A')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=current_row, column=col, value=value)
                cell.style = self.data_style
                
                # Highlight overdue items
                if col == 5 and isinstance(value, int) and value > 0:
                    cell.fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
            
            current_row += 1
    
    def _build_overdue_analysis_sheet(self, worksheet, report_data: Dict, start_row: int):
        """Build overdue training analysis sheet"""
        current_row = start_row
        
        overdue_data = report_data.get('overdue_training', [])
        if not overdue_data:
            worksheet[f'A{current_row}'] = 'No overdue training found'
            return
        
        # Overdue analysis title
        worksheet[f'A{current_row}'] = 'Overdue Training Analysis'
        worksheet[f'A{current_row}'].font = Font(size=14, bold=True)
        current_row += 2
        
        # Headers
        headers = ['Employee', 'Department', 'Training Program', 'Due Date', 'Days Overdue', 'Priority']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=current_row, column=col, value=header)
            cell.style = self.header_style
        
        current_row += 1
        
        # Overdue data
        for item in overdue_data:
            row_data = [
                item.get('employee_name', ''),
                item.get('department', ''),
                item.get('training_program', ''),
                item.get('due_date', ''),
                item.get('days_overdue', 0),
                item.get('priority', 'Medium')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=current_row, column=col, value=value)
                cell.style = self.data_style
                
                # Color code by priority and days overdue
                if col == 5:  # Days overdue
                    if isinstance(value, int):
                        if value > 30:
                            cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                        elif value > 14:
                            cell.fill = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')
                        elif value > 7:
                            cell.fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
            
            current_row += 1
    
    def _build_data_table_sheet(self, worksheet, sheet_config: Dict, report_data: Dict, start_row: int):
        """Build a data table sheet based on configuration"""
        current_row = start_row
        
        data_key = sheet_config.get('data_key', 'data')
        table_data = report_data.get(data_key, [])
        
        if not table_data:
            worksheet[f'A{current_row}'] = f'No data available for {data_key}'
            return
        
        # Convert data to DataFrame for easier handling
        if isinstance(table_data, list) and len(table_data) > 0:
            if isinstance(table_data[0], dict):
                df = pd.DataFrame(table_data)
            else:
                # Assume it's a list of lists with headers
                headers = table_data[0] if table_data else []
                data_rows = table_data[1:] if len(table_data) > 1 else []
                df = pd.DataFrame(data_rows, columns=headers)
            
            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                for col_num, value in enumerate(r, 1):
                    cell = worksheet.cell(row=current_row, column=col_num, value=value)
                    if current_row == start_row:  # Header row
                        cell.style = self.header_style
                    else:
                        cell.style = self.data_style
                current_row += 1
            
            # Create Excel table
            table_range = f"A{start_row}:{get_column_letter(len(df.columns))}{current_row-1}"
            table = Table(displayName="DataTable", ref=table_range)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=True
            )
            worksheet.add_table(table)
    
    def _add_department_chart(self, worksheet, start_row: int, data_count: int):
        """Add a department performance chart"""
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Department Performance Comparison"
        chart.y_axis.title = 'Score (%)'
        chart.x_axis.title = 'Department'
        
        # Data references (assuming department data starts at row start_row - data_count - 3)
        data_start_row = start_row - data_count - 2
        data = Reference(worksheet, min_col=2, min_row=data_start_row, max_row=data_start_row + data_count - 1, max_col=4)
        cats = Reference(worksheet, min_col=1, min_row=data_start_row + 1, max_row=data_start_row + data_count - 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        
        # Position chart
        worksheet.add_chart(chart, f"A{start_row}")
    
    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width